import pandas as pd
import numpy as np
import openpyxl

import chargeback_rpt.vm_data_validator as vc

# paramater passed from ETL.Data.ipynb 
# veeam_import_path=os.path.abspath(r'D:\JupyterCode\Yit\ChargebackReport\data_import')
# source_type='xlsx'
# source_prefix='VeeamOneReport_Sample'
# source_keyname='Sheet1'
# veeam_fields=['FQDN','IP address','Total Backups Size']
#
# backup_field_name='Total Backups Size'
# vm_field_name="VM"
# veeam_field_name="FQDN"
#
# veeam_filepath=f'{veeam_import_path}\\{source_prefix}.{source_type}'
#
# print(veeam_filepath)

# if cannot , it mean vm_name
def set_vm_name(fqdn):
    vm_name=None
    if fqdn is not None:
        fqdn_split=fqdn.split('.')
        if len(fqdn_split)>0:
          vm_name=fqdn_split[0]  
        else:
          return  fqdn
    return vm_name

#def extract_veeam_report_data():
def extract_veeam_report_data(veeam_filepath,datafield_df):

 try:
        print("Extract data file for extracting")
        # Not Allow to change parameter value due to being permanently defined value in column_table_name column in  datafield_mapping table

        backup_param = "backup_size_gb"
        vm_param = "vm_backup"
        # ===============================================================================================

        vm_field_name = "VM"
        backup_field_name = (datafield_df.query('column_table_name==@backup_param')).iloc[0]['field_source_name']
        veeam_field_name = (datafield_df.query('column_table_name==@vm_param')).iloc[0]['field_source_name']
        veeam_fields = [veeam_field_name, backup_field_name]
        veeam_filepath = veeam_filepath

        print("Load excel file")
        wb_obj = openpyxl.load_workbook(veeam_filepath)
        ws_obj = wb_obj.active
        #print(ws_obj.title)


        print("Find the  initial header row   by seaching  which row contain all cells  as veeam_fields")
        row_header = None
        for row in ws_obj.iter_rows():
            cellList = list(row)
            cellHeaderList = [cell for cell in cellList if cell.value in veeam_fields]
            if len(veeam_fields) == len(cellHeaderList):
                row_header = row
                break
        # make all of them easier to manipulate

        if row_header is None:
         raise Exception("Cannot extract data in this report")

        #print(row_header)
        row_header = list(row_header)
        print("List report rows including header")
        row_all = []
        if (len(row_header) > 0):
            # mark  the row starting point  get data
            rowItem_start = row_header[0].row
            print('Backup report start at row no.', rowItem_start)
            # colItem_start=row_header[0].column

            # get every row including header
            for rowItem in ws_obj.iter_rows(min_row=rowItem_start, values_only=True):
                row_all.append(rowItem)
                #print(rowItem)
        print("Create veeam data frame")
        df = None
        if (len(row_all) >= 2):  # header  and at lest 1 item
            # get only header
            colList = row_all[0]
            # print(colList)

            # list data since index 1 onword
            rowList = row_all[1:]
            df = pd.DataFrame(rowList, columns=colList)

            # create df
            df = df[veeam_fields]

            # Change requirment at 20 Aug 21
            # Name column in Veeam report (it is VM Name)
            df[vm_field_name] = df[veeam_field_name]

            # FQDN column in Veeam report xxx.bdms.co.th
            #df[vm_field_name] = df[veeam_field_name].apply(set_vm_name)


            df=df.drop(columns=veeam_field_name)

            df = df.fillna(value=np.nan)
            df=df.dropna( how='all')


            df = df.replace(r'^\s*$', np.NaN, regex=True)

            df[vm_field_name].fillna(method='ffill',inplace=True)

            def convert_any_to_numeric(item):
              try:
                if vc.isnan(item)==False:
                 if  isinstance(item,str)==True:
                    return float(item)
                 else:
                    return item
                else:
                  raise Exception("found some items in file are blank")

              except Exception as error:
                  raise Exception(error)


            # check if dtype is object
            df[backup_field_name]=df[backup_field_name].apply(convert_any_to_numeric)

            df= df.groupby(vm_field_name, as_index=False).sum()

        #print(df)
        return df
 except Exception as error:
    raise Exception(error)



